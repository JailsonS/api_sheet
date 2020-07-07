from app import app
from openpyxl import load_workbook
from flask import jsonify, request
from app.models.auth import Auth

# routes - /excel/info - extract information from excel file
@app.route('/excel/info', methods=['POST', 'GET'])
def excel():

    # check the method
    if request.method == 'POST':
        
        # standard response
        response = {'res_tab':'', 'res':'', 'err':'', 'jwt_token':''}

        # get request data
        email = request.form['email']
        f = request.files['sample_file']
        file_name = f.filename

        # check email and extension file
        if email != '' and file_name.endswith('.xlsx'):
            
            # check the e-mail
            auth = Auth(email)

            # try create jwt            
            if auth.createJwt():

                response['jwt_token'] = str(auth.jwt_token)
                
                # get worksheet data
                wb = load_workbook(f)
                ws = wb[wb.sheetnames[0]] # first ws
                ws_tabs = sorted(wb.sheetnames) # tabs

                # get only values
                ws_values = []
                for row in ws.iter_rows(min_row=1, max_col=3, max_row=10, values_only=True):
                    if row[0] is not None:
                        ws_values.append(row)
                
                # slice data
                keys = ws_values[:1][0]
                values = ws_values[1:]
                data = []

                # get dict
                for item in values:
                    data.append( dict(zip(keys, item)) )
                
                response['res_tab'] = ws_tabs
                response['res'] = data

                return response
            else:
                response['err'] = 'INVALID EMAIL'
                return response
        
        else:
            response['err'] = 'INVALID INPUT DATA'
            return response
    else:
        response['err'] = 'INVALID METHOD'
        return response
    
    

